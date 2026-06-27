import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


ONAM_RECOMMENDATION_PATH = Path("reports/outputs/onam_price_recommendations.xlsx")
ROBUST_RESULTS_PATH = Path("reports/outputs/robust_model_results.xlsx")
FEATURE_IMPORTANCE_PATH = Path("reports/outputs/robust_model_feature_importance.xlsx")

OUTPUT_DIR = Path("reports/outputs")
FINAL_REPORT_PATH = OUTPUT_DIR / "final_munch_onam_pricing_report.xlsx"


def load_required_files():
    missing_files = []

    for path in [
        ONAM_RECOMMENDATION_PATH,
        ROBUST_RESULTS_PATH,
        FEATURE_IMPORTANCE_PATH,
    ]:
        if not path.exists():
            missing_files.append(str(path))

    if missing_files:
        raise FileNotFoundError(
            "Missing required files:\n" + "\n".join(missing_files)
        )


def read_onam_report():
    overall_summary = pd.read_excel(
        ONAM_RECOMMENDATION_PATH,
        sheet_name="Overall_Summary"
    )

    by_source_sheet = pd.read_excel(
        ONAM_RECOMMENDATION_PATH,
        sheet_name="By_Source_Sheet"
    )

    by_market_type = pd.read_excel(
        ONAM_RECOMMENDATION_PATH,
        sheet_name="By_Market_Type"
    )

    top_markets = pd.read_excel(
        ONAM_RECOMMENDATION_PATH,
        sheet_name="Top_Markets"
    )

    detailed_recommendations = pd.read_excel(
        ONAM_RECOMMENDATION_PATH,
        sheet_name="Detailed_Recommendations"
    )

    return {
        "overall_summary": overall_summary,
        "by_source_sheet": by_source_sheet,
        "by_market_type": by_market_type,
        "top_markets": top_markets,
        "detailed_recommendations": detailed_recommendations,
    }


def read_model_outputs():
    robust_results = pd.read_excel(ROBUST_RESULTS_PATH)
    feature_importance = pd.read_excel(FEATURE_IMPORTANCE_PATH)

    return robust_results, feature_importance


def create_executive_summary(overall_summary, robust_results):
    metric_map = dict(
        zip(overall_summary["Metric"], overall_summary["Value"])
    )

    robust_test = robust_results[
        robust_results["Model"].astype(str).str.contains("GradientBoosting_TEST")
    ].copy()

    if robust_test.empty:
        best_test = robust_results[
            robust_results["Model"].astype(str).str.contains("_TEST")
        ].sort_values("MAE").iloc[0]
    else:
        best_test = robust_test.iloc[0]

    recommended_asp = metric_map.get("Predicted ASP Mean")
    conservative_asp = metric_map.get("Conservative ASP Mean")
    aggressive_asp = metric_map.get("Aggressive ASP Mean")
    actual_asp = metric_map.get("Actual ASP Mean")

    uplift = recommended_asp - actual_asp
    uplift_pct = (uplift / actual_asp) * 100 if actual_asp else None

    executive_summary = pd.DataFrame({
        "Section": [
            "Final Recommendation",
            "Pricing Range",
            "Current Actual ASP",
            "Recommended ASP",
            "Recommended Uplift",
            "Model Used",
            "Model MAE",
            "Model RMSE",
            "Model R2",
            "Business Interpretation",
        ],
        "Finding": [
            "Use ₹5.33 per 10g as the central Onam ASP recommendation.",
            f"Recommended range: ₹{conservative_asp:.2f} to ₹{aggressive_asp:.2f} per 10g.",
            f"Latest Onam-relevant actual ASP is ₹{actual_asp:.2f} per 10g.",
            f"Predicted ASP is ₹{recommended_asp:.2f} per 10g.",
            f"Potential uplift is ₹{uplift:.2f} per 10g, equal to {uplift_pct:.2f}%.",
            str(best_test["Model"]),
            round(best_test["MAE"], 4),
            round(best_test["RMSE"], 4),
            round(best_test["R2"], 4),
            "Kerala and high-distribution South markets can support premium pricing near the upper band.",
        ]
    })

    return executive_summary


def create_business_recommendation_table(by_source_sheet, top_markets):
    source_sheet_view = by_source_sheet.copy()

    rename_map = {
        "Source_Sheet": "Market Segment",
        "rows": "Rows",
        "actual_asp_mean": "Actual ASP",
        "predicted_asp_mean": "Recommended ASP",
        "conservative_asp_mean": "Conservative ASP",
        "aggressive_asp_mean": "Aggressive ASP",
        "perk_asp_mean": "Perk ASP",
        "snakker_asp_mean": "Snakker ASP",
        "competitor_avg_asp": "Competitor Avg ASP",
        "munch_vol_share": "Munch Volume Share %",
        "munch_val_share": "Munch Value Share %",
        "munch_distribution_strength": "Munch Distribution Strength",
    }

    source_sheet_view = source_sheet_view.rename(columns=rename_map)

    numeric_cols = source_sheet_view.select_dtypes(include=["float64", "int64"]).columns
    for col in numeric_cols:
        if col != "Rows":
            source_sheet_view[col] = source_sheet_view[col].round(3)

    top_market_view = top_markets.head(25).copy()

    top_market_view = top_market_view.rename(columns={
        "Market_Type": "Market Type",
        "Market": "Market",
        "rows": "Rows",
        "actual_asp_mean": "Actual ASP",
        "predicted_asp_mean": "Recommended ASP",
        "conservative_asp_mean": "Conservative ASP",
        "aggressive_asp_mean": "Aggressive ASP",
        "perk_asp_mean": "Perk ASP",
        "snakker_asp_mean": "Snakker ASP",
        "munch_vol_share": "Munch Volume Share %",
        "munch_val_share": "Munch Value Share %",
        "munch_distribution_strength": "Munch Distribution Strength",
    })

    numeric_cols = top_market_view.select_dtypes(include=["float64", "int64"]).columns
    for col in numeric_cols:
        if col != "Rows":
            top_market_view[col] = top_market_view[col].round(3)

    return source_sheet_view, top_market_view


def create_feature_importance_summary(feature_importance):
    fi = feature_importance.copy()

    fi["Clean Feature"] = (
        fi["Feature"]
        .astype(str)
        .str.replace("num__", "", regex=False)
        .str.replace("cat__", "", regex=False)
    )

    fi = fi[["Clean Feature", "Importance"]].head(30)
    fi["Importance"] = fi["Importance"].round(5)

    return fi


def create_methodology_sheet():
    methodology = pd.DataFrame({
        "Step": [
            1,
            2,
            3,
            4,
            5,
            6,
            7,
        ],
        "Description": [
            "Loaded Nielsen-style Munch pricing dataset from Excel.",
            "Cleaned rows with missing target ASP and removed metadata columns.",
            "Removed leakage columns such as Munch-relative price gaps.",
            "Engineered competitor, distribution, share, market, and Onam-relevance features.",
            "Trained baseline and advanced models using time-based validation.",
            "Selected robust Gradient Boosting model for final recommendation.",
            "Generated Onam ASP recommendation bands using model MAE.",
        ]
    })

    return methodology


def create_assumptions_sheet():
    assumptions = pd.DataFrame({
        "Assumption": [
            "Target variable",
            "Final model",
            "Validation approach",
            "Onam proxy",
            "Pricing band",
            "Leakage handling",
            "Interpretation",
        ],
        "Details": [
            "Munch_ASP_Rs_per_10g",
            "Robust Gradient Boosting model without high-cardinality Period label.",
            "Train on years before 2026 and test on 2026.",
            "South/Kerala market plus Aug/Sep/Q2/Q3 period proxy.",
            "Conservative and aggressive bands are based on model MAE.",
            "Munch Value, Munch Volume, and Munch-derived price gap variables were excluded from final decision model.",
            "Recommended ASP should be used as a pricing decision-support range, not as an exact mandatory price.",
        ]
    })

    return assumptions


def write_final_report(
    executive_summary,
    source_sheet_view,
    top_market_view,
    by_market_type,
    robust_results,
    feature_summary,
    methodology,
    assumptions,
    detailed_recommendations,
):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(FINAL_REPORT_PATH, engine="openpyxl") as writer:
        executive_summary.to_excel(
            writer,
            sheet_name="Executive_Summary",
            index=False
        )

        source_sheet_view.to_excel(
            writer,
            sheet_name="Segment_Recommendation",
            index=False
        )

        top_market_view.to_excel(
            writer,
            sheet_name="Top_Market_Recommendations",
            index=False
        )

        by_market_type.to_excel(
            writer,
            sheet_name="Market_Type_Summary",
            index=False
        )

        robust_results.to_excel(
            writer,
            sheet_name="Model_Performance",
            index=False
        )

        feature_summary.to_excel(
            writer,
            sheet_name="Feature_Importance",
            index=False
        )

        methodology.to_excel(
            writer,
            sheet_name="Methodology",
            index=False
        )

        assumptions.to_excel(
            writer,
            sheet_name="Assumptions",
            index=False
        )

        detailed_recommendations.to_excel(
            writer,
            sheet_name="Detailed_Data",
            index=False
        )


def format_workbook(path):
    wb = load_workbook(path)

    header_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True
    )

    title_fill = PatternFill(
        start_color="D9EAF7",
        end_color="D9EAF7",
        fill_type="solid"
    )

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top")

        for column_cells in ws.columns:
            max_length = 0
            col_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                value = cell.value
                if value is not None:
                    max_length = max(max_length, len(str(value)))

            adjusted_width = min(max_length + 3, 45)
            ws.column_dimensions[col_letter].width = adjusted_width

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "0.000"

        ws.sheet_view.showGridLines = False

    if "Executive_Summary" in wb.sheetnames:
        ws = wb["Executive_Summary"]

        for row in range(2, ws.max_row + 1):
            ws[f"A{row}"].fill = title_fill
            ws[f"A{row}"].font = Font(bold=True)

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 90

    wb.save(path)


def print_final_message():
    print("=" * 80)
    print("FINAL REPORT GENERATED")
    print("=" * 80)
    print(f"Saved to: {FINAL_REPORT_PATH}")
    print("\nMain sheets created:")
    print("- Executive_Summary")
    print("- Segment_Recommendation")
    print("- Top_Market_Recommendations")
    print("- Market_Type_Summary")
    print("- Model_Performance")
    print("- Feature_Importance")
    print("- Methodology")
    print("- Assumptions")
    print("- Detailed_Data")


def main():
    load_required_files()

    onam_data = read_onam_report()
    robust_results, feature_importance = read_model_outputs()

    executive_summary = create_executive_summary(
        onam_data["overall_summary"],
        robust_results
    )

    source_sheet_view, top_market_view = create_business_recommendation_table(
        onam_data["by_source_sheet"],
        onam_data["top_markets"]
    )

    feature_summary = create_feature_importance_summary(feature_importance)
    methodology = create_methodology_sheet()
    assumptions = create_assumptions_sheet()

    write_final_report(
        executive_summary=executive_summary,
        source_sheet_view=source_sheet_view,
        top_market_view=top_market_view,
        by_market_type=onam_data["by_market_type"],
        robust_results=robust_results,
        feature_summary=feature_summary,
        methodology=methodology,
        assumptions=assumptions,
        detailed_recommendations=onam_data["detailed_recommendations"],
    )

    format_workbook(FINAL_REPORT_PATH)

    print_final_message()


if __name__ == "__main__":
    main()